"""
Importa candidatos desde el Excel al SQLite.
Uso: python importar_excel.py [--dry-run]
"""
import os, sys, re
from datetime import datetime, date

EXCEL_PATH = os.path.join(os.path.dirname(__file__),
    "R2 - 2022 - 2023 - Zona Norte Nuevo Formato.xlsx")

sys.path.insert(0, os.path.dirname(__file__))
from app.database import Base, engine, SessionLocal
from app.models.candidato import Candidato
import openpyxl

Base.metadata.create_all(bind=engine)

DRY_RUN = "--dry-run" in sys.argv

# ─────────────────────────────────────────────
# Utilidades de limpieza
# ─────────────────────────────────────────────

def clean_str(val):
    """Convierte a string limpio o None."""
    if val is None:
        return None
    s = str(val).strip()
    # Quitar .0 de números que se leen como float
    if re.match(r'^-?\d+\.0$', s):
        s = s[:-2]
    if s.upper() in ("", "N/A", "NA", "NONE", "-", "0", "0.0"):
        return None
    return s

def clean_fecha(val):
    """Convierte datetime/date/string a 'YYYY-MM-DD' o None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        try:
            return val.strftime("%Y-%m-%d")
        except Exception:
            return None
    s = str(val).strip()
    if not s or s.upper() in ("N/A", "NA", "-"):
        return None
    # Intentar parsear varios formatos
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s.split(" ")[0].split("T")[0], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None

def clean_bool(val):
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("SI", "SÍ", "S", "YES", "1", "TRUE", "X")

def clean_talla(val):
    """Tallas numéricas como '32.0' → '32', strings como 'S' → 'S'."""
    if val is None:
        return None
    s = str(val).strip()
    if re.match(r'^\d+\.0$', s):
        s = str(int(float(s)))
    if s.upper() in ("", "N/A", "NA", "-"):
        return None
    return s

def clean_cedula(val):
    """Cédula sin decimales ni espacios."""
    if val is None:
        return None
    s = str(val).strip().replace(" ", "").replace(".", "").replace(",", "")
    if re.match(r'^\d+\.0+$', s):
        s = str(int(float(val)))
    if not s or s in ("0",):
        return None
    return s

def clean_nombre(val):
    """Nombre limpio — descarta si parece email, número puro o basura."""
    s = clean_str(val)
    if not s:
        return None
    if "@" in s:           # Es un email en la columna nombre
        return None
    if re.match(r'^\d+$', s):  # Solo números
        return None
    return s.title()       # Capitalizar bien

def clean_telefono(val):
    """Teléfono — solo dígitos y guiones."""
    s = clean_str(val)
    if not s:
        return None
    # Quitar todo excepto dígitos, +, -, espacio
    t = re.sub(r'[^\d\+\-\s]', '', s).strip()
    return t if t else None


# ─────────────────────────────────────────────
# Mapa para hojas ODT / SDT-JDT / Part Time
# Columnas 0-indexed (A=0, B=1, ...)
# ─────────────────────────────────────────────
COL_PRINCIPAL = {
    1:  ("observaciones_analistas", clean_str),
    2:  ("reclutador",              clean_str),
    3:  ("zona",                    clean_str),
    4:  ("region",                  clean_str),
    5:  ("departamento",            clean_str),
    6:  ("municipio",               clean_str),
    7:  ("localidad",               clean_str),
    8:  ("ciudad_aplica",           clean_str),
    9:  ("negocio",                 clean_str),
    10: ("cargo",                   clean_str),
    11: ("fecha_nacimiento",        clean_fecha),
    12: ("genero",                  clean_str),
    13: ("tipo_documento",          clean_str),
    14: ("cedula",                  clean_cedula),
    15: ("nombre",                  clean_nombre),
    16: ("telefono_contacto",       clean_telefono),
    17: ("correo",                  clean_str),
    18: ("direccion",               clean_str),
    19: ("talla_pantalon",          clean_talla),
    20: ("talla_camiseta",          clean_talla),
    21: ("talla_zapatos",           clean_talla),
    22: ("fuente",                  clean_str),
    23: ("medio_transporte",        clean_str),
    24: ("fecha_prog_operaciones",  clean_fecha),
    25: ("entrevistador_1",         clean_str),
    26: ("fecha_retro_operaciones", clean_fecha),
    27: ("resultado_operaciones",   clean_str),
    28: ("fecha_prog_rrhh",         clean_fecha),
    29: ("entrevistador_2",         clean_str),
    30: ("fecha_retro_rrhh",        clean_fecha),
    31: ("resultado_rrhh",          clean_str),
    32: ("fecha_envio_emo",         clean_fecha),
    33: ("fecha_recibido_emo",      clean_fecha),
    34: ("concepto_emo",            clean_str),
    35: ("proveedor_emo",           clean_str),
    36: ("comentarios_emo",         clean_str),
    37: ("fecha_envio_es",          clean_fecha),
    38: ("fecha_recibido_es",       clean_fecha),
    39: ("concepto_es",             clean_str),
    40: ("proveedor_es",            clean_str),
    41: ("comentarios_es",          clean_str),
    42: ("fecha_contratacion",      clean_fecha),
    43: ("status",                  clean_str),
    44: ("tipo_status",             clean_str),
    45: ("comentarios_status",      clean_str),
    46: ("correo_agradecimiento",   clean_bool),
    47: ("comentarios_operaciones", clean_str),
    48: ("comentarios_rrhh",        clean_str),
    49: ("comunicacion_candidatos", clean_str),
    50: ("gestion_hello",           clean_bool),
    51: ("facturacion_emo",         clean_str),
    52: ("facturacion_es",          clean_str),
    53: ("lista_negra",             clean_bool),
}

# ─────────────────────────────────────────────
# Mapa para hoja SENA R2
# ─────────────────────────────────────────────
COL_SENA = {
    2:  ("zona",               clean_str),
    3:  ("region",             clean_str),
    # 4: CECO — sin campo en modelo
    5:  ("departamento",       clean_str),
    6:  ("municipio",          clean_str),
    7:  ("localidad",          clean_str),
    8:  ("ciudad_aplica",      clean_str),
    9:  ("fecha_nacimiento",   clean_fecha),
    # 10: EDAD — calculado, omitir
    11: ("genero",             clean_str),
    12: ("cedula",             clean_cedula),
    13: ("nombre",             clean_nombre),
    14: ("telefono_contacto",  clean_telefono),
    15: ("correo",             clean_str),
    16: ("direccion",          clean_str),
    17: ("talla_pantalon",     clean_talla),
    18: ("talla_camiseta",     clean_talla),
    19: ("talla_zapatos",      clean_talla),
    20: ("fecha_envio_emo",    clean_fecha),
    21: ("fecha_recibido_emo", clean_fecha),
    22: ("concepto_emo",       clean_str),
    23: ("proveedor_emo",      clean_str),
    24: ("comentarios_emo",    clean_str),
    25: ("fecha_envio_es",     clean_fecha),
    26: ("fecha_recibido_es",  clean_fecha),
    27: ("concepto_es",        clean_str),
    28: ("proveedor_es",       clean_str),
    29: ("comentarios_es",     clean_str),
    30: ("fecha_contratacion", clean_fecha),
    # 31: BACKUP/OFICIAL — sin campo
    32: ("status",             clean_str),
    33: ("tipo_status",        clean_str),
    34: ("comentarios_status", clean_str),
    35: ("correo_agradecimiento", clean_bool),
    36: ("comunicacion_candidatos", clean_str),
    # 37: Revisión de Aptos — sin campo
    38: ("facturacion_emo",    clean_str),
    39: ("facturacion_es",     clean_str),
}

HOJAS = {
    "ODT R2":       ("ODT — Operador de Tienda",        COL_PRINCIPAL),
    "SDT - JDT R2": ("SDT/JDT — Supervisor/Jefe de Tienda", COL_PRINCIPAL),
    "OP Part Time": ("Part Time",                        COL_PRINCIPAL),
    "SENA R2":      ("SENA",                             COL_SENA),
}


def import_sheet(ws, tipo_formulario, col_map, db, stats):
    insertados = omitidos = errores = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Saltar filas completamente vacías
        if all(v is None or str(v).strip() == "" for v in row):
            omitidos += 1
            continue

        data = {
            "tipo_formulario": tipo_formulario,
            "creado_por": "importacion_excel",
        }

        for col_idx, (campo, fn) in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            data[campo] = fn(val)

        # Cédula y nombre son obligatorios
        if not data.get("cedula") or not data.get("nombre"):
            omitidos += 1
            continue

        # Validar que cédula tenga sentido (al menos 5 dígitos)
        ced = str(data["cedula"])
        if not re.match(r'^\d{5,12}$', ced):
            omitidos += 1
            continue

        try:
            db.add(Candidato(**data))
            insertados += 1
        except Exception as e:
            errores += 1
            print(f"  ERROR fila {row_num}: {e}")
            db.rollback()
            continue

        if insertados % 500 == 0:
            if not DRY_RUN:
                db.commit()
            print(f"    ... {insertados} guardados")

    if not DRY_RUN:
        db.commit()

    stats["insertados"] += insertados
    stats["omitidos"]   += omitidos
    stats["errores"]    += errores
    print(f"  Hoja '{ws.title}': {insertados} insertados | {omitidos} omitidos | {errores} errores")


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: No se encontró:\n  {EXCEL_PATH}")
        sys.exit(1)

    if DRY_RUN:
        print("MODO DRY-RUN - no se guardara nada en la base de datos\n")

    print(f"Abriendo: {os.path.basename(EXCEL_PATH)}\n")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    db = SessionLocal()
    stats = {"insertados": 0, "omitidos": 0, "errores": 0}

    try:
        for nombre_hoja, (tipo, col_map) in HOJAS.items():
            if nombre_hoja not in wb.sheetnames:
                print(f"  (Hoja '{nombre_hoja}' no encontrada, omitida)")
                continue
            print(f"Procesando: {nombre_hoja} | tipo='{tipo}'")
            import_sheet(wb[nombre_hoja], tipo, col_map, db, stats)
    finally:
        db.close()
        wb.close()

    print(f"\n{'='*55}")
    if DRY_RUN:
        print("DRY-RUN completado - NADA fue guardado")
    else:
        print("IMPORTACIÓN COMPLETADA")
    print(f"  Registros insertados : {stats['insertados']}")
    print(f"  Filas omitidas       : {stats['omitidos']}")
    print(f"  Errores              : {stats['errores']}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
