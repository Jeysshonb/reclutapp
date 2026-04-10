"""
Exportación de datos de candidatos a Excel y CSV.
Soporta filtros: negocio, reclutador, status, cargo, zona, region,
                 fecha_desde, fecha_hasta (por fecha_prog_operaciones).
"""
import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.candidato import Candidato, Usuario
from app.routers.auth import get_current_user_download

router = APIRouter(prefix="/export", tags=["exportación"])

COLUMNAS_COMPLETO = [
    # Proceso
    "id", "reclutador", "zona", "region", "negocio", "tipo_formulario", "cargo", "fuente", "ciudad_aplica",
    # Datos personales
    "cedula", "tipo_documento", "nombre", "fecha_nacimiento", "genero",
    "correo", "telefono_contacto", "direccion", "departamento", "municipio", "localidad",
    # Medidas
    "talla_pantalon", "talla_camiseta", "talla_zapatos", "peso", "altura", "imc", "resultado_imc",
    # Inclusión
    "tiene_discapacidad", "tipo_discapacidad", "tiene_certificado_discapacidad",
    # Familia y disponibilidad
    "tiene_hijos", "num_hijos", "edades_hijos", "apoyo_cuidado",
    "familiar_en_ara", "quien_familiar_ara", "medio_transporte",
    "disponibilidad_desplazamiento", "destino_desplazamiento",
    "disponibilidad_reubicacion", "destino_reubicacion",
    # Educación
    "nivel_academico", "anio_graduacion", "titulo",
    "estudia_actualmente", "modalidad_estudio", "proyectos_corto_plazo",
    # Situación laboral
    "situacion_laboral", "tipo_contrato", "salario_actual", "aspiracion_salarial",
    "meses_desempleado", "justificacion_cambio_indefinido",
    # Experiencia 1
    "exp1_empresa", "exp1_cargo", "exp1_actividad", "exp1_ciudad", "exp1_salario",
    "exp1_fecha_inicio", "exp1_fecha_retiro", "exp1_funciones", "exp1_motivo_retiro",
    # Experiencia 2
    "exp2_empresa", "exp2_cargo", "exp2_actividad", "exp2_ciudad", "exp2_salario",
    "exp2_fecha_inicio", "exp2_fecha_retiro", "exp2_funciones", "exp2_motivo_retiro",
    # Experiencia 3
    "exp3_empresa", "exp3_cargo", "exp3_actividad", "exp3_ciudad", "exp3_salario",
    "exp3_fecha_inicio", "exp3_fecha_retiro", "exp3_funciones", "exp3_motivo_retiro",
    # Pipeline
    "fecha_prog_operaciones", "entrevistador_1", "resultado_operaciones", "comentarios_operaciones",
    "fecha_prog_rrhh", "entrevistador_2", "resultado_rrhh", "comentarios_rrhh",
    "concepto_emo", "proveedor_emo", "comentarios_emo",
    "concepto_es", "proveedor_es", "comentarios_es",
    "fecha_contratacion", "status", "tipo_status", "comentarios_status",
    # Auditoría
    "lista_negra", "motivo_lista_negra", "creado_por", "created_at",
]

ENCABEZADOS_COMPLETO = {
    "id": "ID", "reclutador": "Reclutador", "zona": "Zona", "region": "Región",
    "negocio": "Negocio", "tipo_formulario": "Tipo Formulario", "cargo": "Cargo",
    "fuente": "Fuente HV", "ciudad_aplica": "Ciudad Aplica",
    "cedula": "Cédula", "tipo_documento": "Tipo Documento", "nombre": "Nombre",
    "fecha_nacimiento": "Fecha Nacimiento", "genero": "Género",
    "correo": "Correo", "telefono_contacto": "Teléfono", "direccion": "Dirección",
    "departamento": "Departamento", "municipio": "Municipio", "localidad": "Localidad",
    "talla_pantalon": "Talla Pantalón", "talla_camiseta": "Talla Camiseta",
    "talla_zapatos": "Talla Zapatos", "peso": "Peso (kg)", "altura": "Altura (m)",
    "imc": "IMC", "resultado_imc": "Resultado IMC",
    "tiene_discapacidad": "Tiene Discapacidad", "tipo_discapacidad": "Tipo Discapacidad",
    "tiene_certificado_discapacidad": "Cert. Discapacidad",
    "tiene_hijos": "Tiene Hijos", "num_hijos": "Nro Hijos", "edades_hijos": "Edades Hijos",
    "apoyo_cuidado": "Apoyo Cuidado", "familiar_en_ara": "Familiar en Ara",
    "quien_familiar_ara": "Quién Familiar Ara", "medio_transporte": "Transporte",
    "disponibilidad_desplazamiento": "Disponib. Desplazamiento",
    "destino_desplazamiento": "Destino Desplazamiento",
    "disponibilidad_reubicacion": "Disponib. Reubicación",
    "destino_reubicacion": "Destino Reubicación",
    "nivel_academico": "Nivel Académico", "anio_graduacion": "Año Graduación", "titulo": "Título",
    "estudia_actualmente": "Estudia Actualmente", "modalidad_estudio": "Modalidad Estudio",
    "proyectos_corto_plazo": "Proyectos Corto Plazo",
    "situacion_laboral": "Situación Laboral", "tipo_contrato": "Tipo Contrato",
    "salario_actual": "Salario Actual", "aspiracion_salarial": "Aspiración Salarial",
    "meses_desempleado": "Meses Desempleado", "justificacion_cambio_indefinido": "Justif. Cambio",
    "exp1_empresa": "Exp1 Empresa", "exp1_cargo": "Exp1 Cargo", "exp1_actividad": "Exp1 Actividad",
    "exp1_ciudad": "Exp1 Ciudad", "exp1_salario": "Exp1 Salario",
    "exp1_fecha_inicio": "Exp1 Inicio", "exp1_fecha_retiro": "Exp1 Retiro",
    "exp1_funciones": "Exp1 Funciones", "exp1_motivo_retiro": "Exp1 Motivo Retiro",
    "exp2_empresa": "Exp2 Empresa", "exp2_cargo": "Exp2 Cargo", "exp2_actividad": "Exp2 Actividad",
    "exp2_ciudad": "Exp2 Ciudad", "exp2_salario": "Exp2 Salario",
    "exp2_fecha_inicio": "Exp2 Inicio", "exp2_fecha_retiro": "Exp2 Retiro",
    "exp2_funciones": "Exp2 Funciones", "exp2_motivo_retiro": "Exp2 Motivo Retiro",
    "exp3_empresa": "Exp3 Empresa", "exp3_cargo": "Exp3 Cargo", "exp3_actividad": "Exp3 Actividad",
    "exp3_ciudad": "Exp3 Ciudad", "exp3_salario": "Exp3 Salario",
    "exp3_fecha_inicio": "Exp3 Inicio", "exp3_fecha_retiro": "Exp3 Retiro",
    "exp3_funciones": "Exp3 Funciones", "exp3_motivo_retiro": "Exp3 Motivo Retiro",
    "fecha_prog_operaciones": "Fecha Entrevista Ops", "entrevistador_1": "Entrevistador 1",
    "resultado_operaciones": "Resultado Ops", "comentarios_operaciones": "Comentarios Ops",
    "fecha_prog_rrhh": "Fecha Entrevista RRHH", "entrevistador_2": "Entrevistador 2",
    "resultado_rrhh": "Resultado RRHH", "comentarios_rrhh": "Comentarios RRHH",
    "concepto_emo": "Concepto EMO", "proveedor_emo": "Proveedor EMO", "comentarios_emo": "Comentarios EMO",
    "concepto_es": "Concepto ES", "proveedor_es": "Proveedor ES", "comentarios_es": "Comentarios ES",
    "fecha_contratacion": "Fecha Contratación", "status": "Status", "tipo_status": "Tipo Status",
    "comentarios_status": "Comentarios Status",
    "lista_negra": "Lista Negra", "motivo_lista_negra": "Motivo Lista Negra",
    "creado_por": "Creado Por", "created_at": "Fecha Registro",
}

COLUMNAS_EXPORT = [
    "id", "reclutador", "zona", "region", "negocio", "cargo", "fuente",
    "cedula", "tipo_documento", "nombre", "fecha_nacimiento", "genero",
    "correo", "telefono_contacto", "direccion", "departamento", "municipio",
    "ciudad_aplica", "talla_pantalon", "talla_camiseta", "talla_zapatos",
    "peso", "altura", "imc", "resultado_imc", "medio_transporte",
    "nivel_academico", "situacion_laboral", "salario_actual", "aspiracion_salarial",
    "fecha_prog_operaciones", "entrevistador_1", "resultado_operaciones", "comentarios_operaciones",
    "fecha_prog_rrhh", "entrevistador_2", "resultado_rrhh", "comentarios_rrhh",
    "concepto_emo", "proveedor_emo", "comentarios_emo",
    "concepto_es", "proveedor_es", "comentarios_es",
    "fecha_contratacion", "status", "tipo_status", "comentarios_status",
    "lista_negra", "motivo_lista_negra",
    "creado_por", "created_at",
]

ENCABEZADOS = {
    "id": "ID", "reclutador": "Reclutador", "zona": "Zona", "region": "Región",
    "negocio": "Negocio", "cargo": "Cargo", "fuente": "Fuente HV",
    "cedula": "Cédula", "tipo_documento": "Tipo Documento", "nombre": "Nombre Candidato",
    "fecha_nacimiento": "Fecha Nacimiento", "genero": "Género",
    "correo": "Correo", "telefono_contacto": "Teléfono", "direccion": "Dirección",
    "departamento": "Departamento", "municipio": "Municipio", "ciudad_aplica": "Ciudad Aplica",
    "talla_pantalon": "Talla Pantalón", "talla_camiseta": "Talla Camiseta", "talla_zapatos": "Talla Zapatos",
    "peso": "Peso (kg)", "altura": "Altura (m)", "imc": "IMC", "resultado_imc": "Resultado IMC",
    "medio_transporte": "Transporte",
    "nivel_academico": "Nivel Académico", "situacion_laboral": "Situación Laboral",
    "salario_actual": "Salario Actual", "aspiracion_salarial": "Aspiración Salarial",
    "fecha_prog_operaciones": "Fecha Entrevista Ops", "entrevistador_1": "Entrevistador 1",
    "resultado_operaciones": "Resultado Operaciones", "comentarios_operaciones": "Comentarios Ops",
    "fecha_prog_rrhh": "Fecha Entrevista RRHH", "entrevistador_2": "Entrevistador 2",
    "resultado_rrhh": "Resultado RRHH", "comentarios_rrhh": "Comentarios RRHH",
    "concepto_emo": "Concepto EMO", "proveedor_emo": "Proveedor EMO", "comentarios_emo": "Comentarios EMO",
    "concepto_es": "Concepto ES", "proveedor_es": "Proveedor ES", "comentarios_es": "Comentarios ES",
    "fecha_contratacion": "Fecha Contratación", "status": "Status", "tipo_status": "Tipo Status",
    "comentarios_status": "Comentarios Status",
    "lista_negra": "Lista Negra", "motivo_lista_negra": "Motivo Lista Negra",
    "creado_por": "Creado Por", "created_at": "Fecha Registro",
}


def _get_filtered(
    db: Session,
    negocio: Optional[str],
    reclutador: Optional[str],
    status: Optional[str],
    cargo: Optional[str],
    zona: Optional[str],
    region: Optional[str],
    fecha_desde: Optional[str],
    fecha_hasta: Optional[str],
):
    q = db.query(Candidato).filter(Candidato.deleted_at.is_(None))
    if negocio:
        q = q.filter(Candidato.negocio == negocio)
    if reclutador:
        q = q.filter(Candidato.reclutador == reclutador)
    if status:
        q = q.filter(Candidato.status == status)
    if cargo:
        q = q.filter(Candidato.cargo == cargo)
    if zona:
        q = q.filter(Candidato.zona == zona)
    if region:
        q = q.filter(Candidato.region == region)
    if fecha_desde:
        q = q.filter(Candidato.created_at >= fecha_desde)
    if fecha_hasta:
        q = q.filter(Candidato.created_at <= fecha_hasta + " 23:59:59")
    return q.order_by(Candidato.created_at.desc()).all()


@router.get("/excel")
def exportar_excel(
    negocio: Optional[str] = Query(None),
    reclutador: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    cargo: Optional[str] = Query(None),
    zona: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user_download),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    candidatos = _get_filtered(db, negocio, reclutador, status, cargo, zona, region, fecha_desde, fecha_hasta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    header_fill = PatternFill("solid", fgColor="FE5000")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_i, col_name in enumerate(COLUMNAS_EXPORT, 1):
        cell = ws.cell(row=1, column=col_i, value=ENCABEZADOS.get(col_name, col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    FECHA_COLS = {col for col in COLUMNAS_EXPORT if "fecha" in col or col == "created_at"}

    for row_i, c in enumerate(candidatos, 2):
        for col_i, col_name in enumerate(COLUMNAS_EXPORT, 1):
            val = getattr(c, col_name, None)
            if val is not None and col_name in FECHA_COLS:
                # Convertir a DD/MM/AAAA
                s = str(val).split("T")[0].split(" ")[0]
                parts = s.split("-")
                val = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else s
            elif hasattr(val, "isoformat"):
                val = str(val)
            ws.cell(row=row_i, column=col_i, value=val)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="candidatos_reclutapp.xlsx"'},
    )


@router.get("/csv")
def exportar_csv(
    negocio: Optional[str] = Query(None),
    reclutador: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    cargo: Optional[str] = Query(None),
    zona: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user_download),
):
    import csv

    candidatos = _get_filtered(db, negocio, reclutador, status, cargo, zona, region, fecha_desde, fecha_hasta)

    buf = io.StringIO()
    # Separador ";" — Excel español abre directo sin asistente de importación
    writer = csv.DictWriter(buf, fieldnames=COLUMNAS_EXPORT, extrasaction="ignore", delimiter=";")
    writer.writeheader()
    FECHA_COLS = {col for col in COLUMNAS_EXPORT if "fecha" in col or col == "created_at"}

    for c in candidatos:
        row = {}
        for col in COLUMNAS_EXPORT:
            val = getattr(c, col, None)
            if val is not None and col in FECHA_COLS:
                s = str(val).split("T")[0].split(" ")[0]
                parts = s.split("-")
                val = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else s
            elif hasattr(val, "isoformat"):
                val = str(val)
            row[col] = val
        writer.writerow(row)

    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="candidatos_reclutapp.csv"'},
    )


@router.get("/csv-completo")
def exportar_csv_completo(
    negocio: Optional[str] = Query(None),
    reclutador: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    cargo: Optional[str] = Query(None),
    zona: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user_download),
):
    """CSV con TODOS los campos: datos personales, medidas, inclusión, familia,
    educación, situación laboral, 3 experiencias y pipeline completo."""
    import csv

    candidatos = _get_filtered(db, negocio, reclutador, status, cargo, zona, region, fecha_desde, fecha_hasta)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNAS_COMPLETO, extrasaction="ignore", delimiter=";")
    writer.writeheader()
    FECHA_COLS = {col for col in COLUMNAS_COMPLETO if "fecha" in col or col == "created_at"}
    BOOL_COLS = {col for col in COLUMNAS_COMPLETO if col.startswith("tiene_") or col.startswith("disponibilidad_") or col.startswith("estudia_") or col.startswith("familiar_") or col in ("lista_negra",)}

    for c in candidatos:
        row = {}
        for col in COLUMNAS_COMPLETO:
            val = getattr(c, col, None)
            if val is not None and col in FECHA_COLS:
                s = str(val).split("T")[0].split(" ")[0]
                parts = s.split("-")
                val = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else s
            elif col in BOOL_COLS:
                val = "Sí" if val else "No"
            elif hasattr(val, "isoformat"):
                val = str(val)
            row[col] = val
        writer.writerow(row)

    today = date.today().strftime("%Y%m%d")
    fname = f"fichas_completas_{today}.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/excel-completo")
def exportar_excel_completo(
    negocio: Optional[str] = Query(None),
    reclutador: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    cargo: Optional[str] = Query(None),
    zona: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user_download),
):
    """Excel .xlsx con TODOS los campos — abre directo en Excel sin conversión."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    candidatos = _get_filtered(db, negocio, reclutador, status, cargo, zona, region, fecha_desde, fecha_hasta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requerimientos"

    header_fill = PatternFill("solid", fgColor="FE5000")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", wrap_text=False)

    for col_i, col_name in enumerate(COLUMNAS_COMPLETO, 1):
        cell = ws.cell(row=1, column=col_i, value=ENCABEZADOS_COMPLETO.get(col_name, col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    FECHA_COLS = {col for col in COLUMNAS_COMPLETO if "fecha" in col or col == "created_at"}
    BOOL_COLS = {col for col in COLUMNAS_COMPLETO if col.startswith("tiene_") or col.startswith("disponibilidad_") or col.startswith("estudia_") or col.startswith("familiar_") or col in ("lista_negra",)}

    for row_i, c in enumerate(candidatos, 2):
        for col_i, col_name in enumerate(COLUMNAS_COMPLETO, 1):
            val = getattr(c, col_name, None)
            if val is not None and col_name in FECHA_COLS:
                s = str(val).split("T")[0].split(" ")[0]
                parts = s.split("-")
                val = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else s
            elif col_name in BOOL_COLS:
                val = "Sí" if val else "No"
            elif hasattr(val, "isoformat"):
                val = str(val)
            ws.cell(row=row_i, column=col_i, value=val)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Freeze primera fila
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%Y%m%d")
    fname = f"requerimientos_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
